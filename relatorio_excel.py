"""
Relatorio executivo em Excel - Casa Verde Distribuidora

Gera o arquivo que e entregue ao cliente. Todas as tabelas saem do
pipeline, e os totais sao formulas do Excel, nao valores colados, para
que a planilha continue viva quando o cliente filtrar ou editar.

Uso:
    python src/relatorio_excel.py
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from limpeza import executar_limpeza
from analise import gerar_analises, principais_achados

PASTA_SAIDA = Path("data/output")
ARQUIVO = "relatorio_margem_casa_verde.xlsx"

FONTE = "Arial"

# Paleta sobria, adequada a relatorio executivo
AZUL_ESCURO = "1F3864"
CINZA_CLARO = "F2F2F2"
VERMELHO = "C00000"
VERDE = "375623"
AMBAR = "BF8F00"

MOEDA = 'R$ #,##0.00;[Red]-R$ #,##0.00;"-"'
MOEDA_INT = 'R$ #,##0;[Red]-R$ #,##0;"-"'
PERCENTUAL = '0.0"%";[Red]-0.0"%";"-"'
INTEIRO = '#,##0;-#,##0;"-"'

BORDA_FINA = Border(bottom=Side(style="thin", color="BFBFBF"))

# Rotulos amigaveis. O cliente nao deve ver nome de coluna de banco.
ROTULOS = {
    "categoria": "Categoria",
    "codigo_produto": "Codigo",
    "descricao": "Produto",
    "codigo_cliente": "Codigo",
    "razao_social": "Cliente",
    "segmento": "Segmento",
    "regiao": "Regiao",
    "receita": "Receita",
    "margem": "Margem de contribuicao",
    "margem_pct": "Margem %",
    "quantidade": "Quantidade",
    "desconto_medio": "Desconto medio %",
    "pedidos": "Pedidos",
    "ticket_medio": "Ticket medio",
    "situacao": "Situacao",
    "classe_abc": "Classe ABC",
    "receita_acumulada_pct": "Receita acumulada %",
    "participacao_receita_pct": "Participacao %",
    "custo_frete": "Custo de entrega",
    "custo_entrega": "Custo de entrega",
    "frete_cobrado": "Frete cobrado",
    "frete_subsidiado": "Frete subsidiado",
    "subsidio": "Frete subsidiado",
    "subsidio_sobre_receita_pct": "Subsidio sobre receita %",
    "margem_sem_subsidio": "Margem sem o subsidio",
    "ano_mes": "Mes",
    "faixa_desconto": "Faixa de desconto",
    "linhas": "Itens vendidos",
    "etapa": "Etapa",
    "problema": "Problema encontrado",
    "registros": "Registros",
    "acao": "Acao tomada",
}

COLUNAS_MOEDA = {
    "receita", "margem", "ticket_medio", "custo_frete", "custo_entrega",
    "frete_cobrado", "frete_subsidiado", "subsidio", "margem_sem_subsidio",
}
COLUNAS_PERCENTUAL = {
    "margem_pct", "desconto_medio", "participacao_receita_pct",
    "receita_acumulada_pct", "subsidio_sobre_receita_pct",
}
COLUNAS_INTEIRO = {"quantidade", "pedidos", "linhas", "registros"}

COR_SITUACAO = {
    "PREJUIZO": VERMELHO,
    "CRITICA": VERMELHO,
    "ATENCAO": AMBAR,
    "SAUDAVEL": VERDE,
    "SEM CUSTO": "808080",
}


# ---------------------------------------------------------------------
# Escrita de tabelas
# ---------------------------------------------------------------------

def _largura_coluna(serie, titulo):
    maior = max([len(str(v)) for v in serie.head(200)] + [len(titulo)])
    return min(max(maior + 3, 12), 42)


def escrever_tabela(ws, df, linha_inicial=1, nome_tabela=None, totalizar=None):
    """
    Escreve um DataFrame como tabela formatada e devolve a ultima linha.

    O parametro totalizar recebe as colunas que ganham linha de total,
    escrita como formula SUM sobre o intervalo, nunca como valor.
    """
    colunas = list(df.columns)
    totalizar = totalizar or []

    for j, coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_inicial, column=j, value=ROTULOS.get(coluna, coluna))
        celula.font = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, registro) in enumerate(df.iterrows(), start=1):
        linha = linha_inicial + i
        for j, coluna in enumerate(colunas, start=1):
            valor = registro[coluna]
            if pd.isna(valor):
                valor = None
            celula = ws.cell(row=linha, column=j, value=valor)
            celula.font = Font(name=FONTE, size=10)
            celula.border = BORDA_FINA

            if coluna in COLUNAS_MOEDA:
                celula.number_format = MOEDA
            elif coluna in COLUNAS_PERCENTUAL:
                celula.number_format = PERCENTUAL
            elif coluna in COLUNAS_INTEIRO:
                celula.number_format = INTEIRO

            if coluna == "situacao" and valor in COR_SITUACAO:
                celula.font = Font(name=FONTE, size=10, bold=True, color=COR_SITUACAO[valor])
                celula.alignment = Alignment(horizontal="center")

            if coluna == "classe_abc":
                celula.alignment = Alignment(horizontal="center")

        if i % 2 == 0:
            for j in range(1, len(colunas) + 1):
                ws.cell(row=linha, column=j).fill = PatternFill("solid", fgColor=CINZA_CLARO)

    ultima_linha = linha_inicial + len(df)

    if totalizar:
        linha_total = ultima_linha + 1
        ws.cell(row=linha_total, column=1, value="TOTAL").font = Font(
            name=FONTE, size=10, bold=True
        )
        for j, coluna in enumerate(colunas, start=1):
            if coluna not in totalizar:
                continue
            letra = get_column_letter(j)
            celula = ws.cell(row=linha_total, column=j)
            celula.value = f"=SUM({letra}{linha_inicial + 1}:{letra}{ultima_linha})"
            celula.font = Font(name=FONTE, size=10, bold=True)
            celula.number_format = MOEDA if coluna in COLUNAS_MOEDA else INTEIRO
            celula.border = Border(top=Side(style="thin", color="000000"))
        ultima_linha = linha_total

    for j, coluna in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(j)].width = _largura_coluna(
            df[coluna], ROTULOS.get(coluna, coluna)
        )

    if nome_tabela and len(df) > 0:
        referencia = (
            f"A{linha_inicial}:"
            f"{get_column_letter(len(colunas))}{linha_inicial + len(df)}"
        )
        tabela = Table(displayName=nome_tabela, ref=referencia)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=False
        )
        ws.add_table(tabela)

    ws.freeze_panes = ws.cell(row=linha_inicial + 1, column=1)
    return ultima_linha


def cabecalho_aba(ws, titulo, subtitulo=None):
    ws["A1"] = titulo
    ws["A1"].font = Font(name=FONTE, size=14, bold=True, color=AZUL_ESCURO)
    if subtitulo:
        ws["A2"] = subtitulo
        ws["A2"].font = Font(name=FONTE, size=10, italic=True, color="595959")
    ws.sheet_view.showGridLines = False
    return 4 if subtitulo else 3


# ---------------------------------------------------------------------
# Aba de resumo
# ---------------------------------------------------------------------

def montar_resumo(ws, achados, periodo):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22

    ws["B2"] = "Casa Verde Distribuidora"
    ws["B2"].font = Font(name=FONTE, size=18, bold=True, color=AZUL_ESCURO)

    ws["B3"] = "Analise de margem de contribuicao por produto, cliente e categoria"
    ws["B3"].font = Font(name=FONTE, size=11, color="595959")

    ws["B4"] = f"Periodo analisado: {periodo}"
    ws["B4"].font = Font(name=FONTE, size=10, italic=True, color="595959")

    linha = 6
    ws[f"B{linha}"] = "RESULTADO DO PERIODO"
    ws[f"B{linha}"].font = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
    ws[f"B{linha}"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws[f"C{linha}"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)

    # Os indicadores referenciam a aba de categorias por formula, entao a
    # capa recalcula sozinha se aquela tabela for alterada.
    indicadores = [
        ("Receita total", "=SUM(Categorias!B5:B9)", MOEDA_INT),
        ("Margem de contribuicao", "=SUM(Categorias!C5:C9)", MOEDA_INT),
        ("Margem sobre a receita", "=IFERROR(C8/C7*100,0)", PERCENTUAL),
        ("Frete subsidiado no ano", "=SUM(Categorias!E5:E9)", MOEDA_INT),
    ]

    linha += 1
    for rotulo, formula, formato in indicadores:
        ws[f"B{linha}"] = rotulo
        ws[f"B{linha}"].font = Font(name=FONTE, size=10)
        ws[f"C{linha}"] = formula
        ws[f"C{linha}"].font = Font(name=FONTE, size=11, bold=True)
        ws[f"C{linha}"].number_format = formato
        ws[f"C{linha}"].alignment = Alignment(horizontal="right")
        ws[f"B{linha}"].border = BORDA_FINA
        ws[f"C{linha}"].border = BORDA_FINA
        linha += 1

    linha += 1
    ws[f"B{linha}"] = "PRINCIPAIS ACHADOS"
    ws[f"B{linha}"].font = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
    ws[f"B{linha}"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws[f"C{linha}"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    linha += 1

    categorias = ", ".join(achados["categorias_no_prejuizo"]) or "nenhuma"

    achados_texto = [
        (
            f"A categoria {categorias} opera com margem negativa",
            achados["prejuizo_categorias"],
            MOEDA_INT,
        ),
        (
            f"{achados['produtos_no_prejuizo']} produtos vendem com margem negativa",
            achados["receita_produtos_prejuizo"],
            MOEDA_INT,
        ),
        (
            "Custo de entrega absorvido pela politica de frete gratis",
            achados["frete_subsidiado_ano"],
            MOEDA_INT,
        ),
    ]

    for texto, valor, formato in achados_texto:
        ws[f"B{linha}"] = texto
        ws[f"B{linha}"].font = Font(name=FONTE, size=10)
        ws[f"B{linha}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{linha}"] = float(valor)
        ws[f"C{linha}"].font = Font(name=FONTE, size=11, bold=True, color=VERMELHO)
        ws[f"C{linha}"].number_format = formato
        ws[f"C{linha}"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[linha].height = 22
        ws[f"B{linha}"].border = BORDA_FINA
        ws[f"C{linha}"].border = BORDA_FINA
        linha += 1

    linha += 1
    ws[f"B{linha}"] = "COMO LER ESTE RELATORIO"
    ws[f"B{linha}"].font = Font(name=FONTE, size=11, bold=True, color=AZUL_ESCURO)
    linha += 1

    notas = [
        "Margem de contribuicao desconta do faturamento o custo da mercadoria, "
        "o custo real da entrega e a comissao do vendedor. E diferente da margem "
        "bruta mostrada pelo sistema, que considera apenas mercadoria.",
        "Frete subsidiado e a parcela do custo de entrega que a empresa absorveu "
        "e nao cobrou do cliente.",
        "Classe ABC segue o criterio de receita acumulada. Classe A concentra os "
        "primeiros 80 por cento do faturamento.",
        "Produtos sem custo de compra cadastrado aparecem no faturamento, mas "
        "ficam fora do calculo de margem. A aba Qualidade dos dados lista quantos sao.",
    ]

    for nota in notas:
        ws[f"B{linha}"] = nota
        ws[f"B{linha}"].font = Font(name=FONTE, size=9, color="404040")
        ws[f"B{linha}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[linha].height = 42
        linha += 1

    linha += 1
    ws[f"B{linha}"] = (
        "Dados sinteticos gerados para demonstracao metodologica. "
        "Os valores nao representam empresa real."
    )
    ws[f"B{linha}"].font = Font(name=FONTE, size=8, italic=True, color="808080")


# ---------------------------------------------------------------------
# Montagem do arquivo
# ---------------------------------------------------------------------

def gerar_relatorio(df, relatorio_qualidade, caminho=None):
    analises = gerar_analises(df)
    achados = principais_achados(df)

    periodo = (
        f"{df['data_pedido'].min():%d/%m/%Y} a {df['data_pedido'].max():%d/%m/%Y}"
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo"
    montar_resumo(ws, achados, periodo)

    abas = [
        (
            "Categorias",
            "Resultado por categoria",
            "Ordenado da pior para a melhor margem",
            analises["resumo_categoria"][[
                "categoria", "receita", "margem", "margem_pct",
                "frete_subsidiado", "participacao_receita_pct", "pedidos", "situacao",
            ]],
            ["receita", "margem", "frete_subsidiado", "pedidos"],
            "TabelaCategorias",
        ),
        (
            "Produtos no prejuizo",
            "Produtos que vendem e dao prejuizo",
            "Apenas itens com receita anual acima de mil reais",
            analises["produtos_prejuizo"][[
                "codigo_produto", "descricao", "categoria", "receita",
                "margem", "margem_pct", "quantidade", "desconto_medio", "pedidos",
            ]],
            ["receita", "margem", "quantidade"],
            "TabelaPrejuizo",
        ),
        (
            "Ranking de produtos",
            "Todos os produtos com curva ABC",
            "Classe A concentra os primeiros 80 por cento da receita",
            analises["ranking_produtos"][[
                "codigo_produto", "descricao", "categoria", "classe_abc",
                "receita", "receita_acumulada_pct", "margem", "margem_pct",
                "quantidade", "desconto_medio", "situacao",
            ]],
            ["receita", "margem", "quantidade"],
            "TabelaProdutos",
        ),
        (
            "Clientes",
            "Carteira de clientes por margem",
            "Cliente classe A com margem baixa e o caso mais critico da carteira",
            analises["ranking_clientes"][[
                "codigo_cliente", "razao_social", "segmento", "regiao",
                "classe_abc", "receita", "margem", "margem_pct",
                "ticket_medio", "desconto_medio", "frete_subsidiado", "pedidos",
            ]],
            ["receita", "margem", "frete_subsidiado", "pedidos"],
            "TabelaClientes",
        ),
        (
            "Frete",
            "Custo da politica de frete gratis",
            "A ultima coluna mostra a margem que existiria sem o subsidio",
            analises["custo_frete_gratis"][[
                "categoria", "receita", "custo_entrega", "frete_cobrado",
                "subsidio", "subsidio_sobre_receita_pct", "margem", "margem_sem_subsidio",
            ]],
            ["receita", "custo_entrega", "frete_cobrado", "subsidio", "margem"],
            "TabelaFrete",
        ),
        (
            "Desconto",
            "Margem por faixa de desconto concedido",
            "Identifica o ponto em que o desconto passa a destruir resultado",
            analises["impacto_desconto"],
            ["receita", "margem", "linhas"],
            "TabelaDesconto",
        ),
        (
            "Evolucao mensal",
            "Receita e margem mes a mes",
            None,
            analises["evolucao_mensal"],
            ["receita", "margem", "pedidos"],
            "TabelaEvolucao",
        ),
        (
            "Qualidade dos dados",
            "Correcoes aplicadas na base recebida",
            "Nenhum registro foi alterado sem constar nesta lista",
            relatorio_qualidade.como_dataframe(),
            None,
            "TabelaQualidade",
        ),
    ]

    for titulo_aba, titulo, subtitulo, dados, totais, nome_tabela in abas:
        ws = wb.create_sheet(titulo_aba[:31])
        inicio = cabecalho_aba(ws, titulo, subtitulo)
        escrever_tabela(ws, dados, linha_inicial=inicio,
                        nome_tabela=nome_tabela, totalizar=totais)

    caminho = Path(caminho) if caminho else PASTA_SAIDA / ARQUIVO
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho


if __name__ == "__main__":
    df, rel = executar_limpeza(salvar=False)
    destino = gerar_relatorio(df, rel)
    print(f"Relatorio gravado em {destino}")
